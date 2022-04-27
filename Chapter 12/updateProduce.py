#! python3
# updateProduce.py - Correct costs in produce sales spreasdsheet.

import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\TC\Desktop\produceSales.xlsx')
sheet = wb['Sheet']

#The produce types and their updated prices.
PRICE_UPDATES =     {'Garlic': 200,
                     'Celery': 100,
                     'Lemon': 400}

#Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row): #skip the first row
    produceName = sheet.cell(row=rowNum, column = 1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updateProduceSales.xlsx')